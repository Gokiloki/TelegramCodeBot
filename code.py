import telebot
from telebot import types
from openpyxl import load_workbook
import openpyxl
import os
from datetime import datetime

bot = telebot.TeleBot('''токен бота''')
admin_password = 'admin'
into_password = 'proba'
authorized_users = {}


def init_users_db():
    if not os.path.exists("users.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["chat_id", "username", "first_name", "last_name", "date_joined", "authorized"])
        wb.save("users.xlsx")


def save_user(message, is_authorized=False):
    try:
        wb = load_workbook("users.xlsx")
        ws = wb.active

        user_exists = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == message.chat.id:
                user_exists = True
                if is_authorized:
                    ws.cell(row=row, column=6).value = "Yes"
                break

        if not user_exists:
            ws.append([
                message.chat.id,
                message.chat.username,
                message.chat.first_name,
                message.chat.last_name,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Yes" if is_authorized else "No"
            ])

        wb.save("users.xlsx")
        wb.close()
    except Exception as e:
        print(f"Ошибка при сохранении пользователя: {e}")


def check_user_authorization(user_id):
    try:
        if not os.path.exists("users.xlsx"):
            return False

        wb = load_workbook("users.xlsx")
        ws = wb.active

        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == user_id:
                auth_status = ws.cell(row=row, column=6).value
                wb.close()
                return auth_status == "Yes"

        wb.close()
        return False
    except Exception as e:
        print(f"Ошибка при проверке авторизации: {e}")
        return False


def auth_required(func):
    def wrapper(message):
        user_id = message.from_user.id

        if user_id not in authorized_users and not check_user_authorization(user_id):
            bot.send_message(message.chat.id, "❌ Доступ запрещен. Сначала авторизуйтесь через /start")
            return

        if user_id not in authorized_users:
            authorized_users[user_id] = True

        return func(message)

    return wrapper


def broadcast_message(text):
    try:
        wb = load_workbook("users.xlsx")
        ws = wb.active

        success_count = 0
        fail_count = 0

        for row in range(2, ws.max_row + 1):
            chat_id = ws.cell(row=row, column=1).value
            if chat_id:
                try:
                    bot.send_message(chat_id, text)
                    success_count += 1
                except Exception as e:
                    print(f"Ошибка отправки для {chat_id}: {e}")
                    fail_count += 1

        wb.close()
        return success_count, fail_count
    except Exception as e:
        print(f"Ошибка при рассылке: {e}")
        return 0, 0


def request_password(message):
    msg = bot.send_message(message.chat.id,
                           "🔐 <b>Требуется авторизация</b>\n\n"
                           "Для использования бота введите пароль:",
                           parse_mode='HTML')
    bot.register_next_step_handler(msg, check_initial_password)


def check_initial_password(message):
    user_id = message.from_user.id

    if message.text == into_password:
        authorized_users[user_id] = True
        save_user(message, is_authorized=True)

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton('📋 Информация')
        btn2 = types.KeyboardButton('⚙️ Админ панель')
        btn3 = types.KeyboardButton('🔍 Поиск по номеру заказа')
        markup.row(btn1, btn2)
        markup.row(btn3)

        bot.send_message(message.chat.id,
                         "✅ <b>Авторизация успешна!</b>\n\n"
                         "Добро пожаловать в систему управления заказами Горно-Обогатительного Комбината.",
                         parse_mode='HTML',
                         reply_markup=markup)
    else:
        msg = bot.send_message(message.chat.id,
                               "❌ <b>Неверный пароль</b>\n\n"
                               "Попробуйте еще раз:",
                               parse_mode='HTML')
        bot.register_next_step_handler(msg, check_initial_password)


def secure(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton('📋 Информация')
    btn2 = types.KeyboardButton('⚙️ Админ панель')
    btn3 = types.KeyboardButton('🔍 Поиск по номеру заказа')
    markup.row(btn1, btn2)
    markup.row(btn3)

    if message.text == admin_password:
        admin_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        admin_btn1 = types.KeyboardButton('📊 Загрузить таблицу')
        admin_btn2 = types.KeyboardButton('📢 Сделать рассылку')
        admin_btn3 = types.KeyboardButton('📨 Написать в чат')
        admin_btn4 = types.KeyboardButton('↩️ Назад в главное меню')
        admin_markup.row(admin_btn1, admin_btn2)
        admin_markup.row(admin_btn3, admin_btn4)

        bot.send_message(message.chat.id, '✅ Авторизация успешна. Доступны функции админ панели.',
                         reply_markup=admin_markup)
    else:
        bot.send_message(message.chat.id, '❌ Неверный пароль. Доступ ограничен.', reply_markup=markup)


def admin_table_add(message):
    try:
        if message.document and message.document.mime_type in [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel']:

            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            file_path = 'Table.xlsx'

            with open(file_path, 'wb') as new_file:
                new_file.write(downloaded_file)

            try:
                wb = load_workbook(file_path)
                ws = wb.active

                admin_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                admin_btn1 = types.KeyboardButton('📊 Загрузить таблицу')
                admin_btn2 = types.KeyboardButton('📢 Сделать рассылку')
                admin_btn3 = types.KeyboardButton('📨 Написать в чат')
                admin_btn4 = types.KeyboardButton('↩️ Назад в главное меню')
                admin_markup.row(admin_btn1, admin_btn2)
                admin_markup.row(admin_btn3, admin_btn4)

                bot.reply_to(message, '✅ Файл таблицы успешно загружен и проверен.', reply_markup=admin_markup)
            except Exception as e:
                bot.reply_to(message, f"❌ Файл содержит ошибки или имеет неверный формат: {e}")
        else:
            bot.reply_to(message, "❌ Требуется файл в формате Excel (.xlsx)")
    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка обработки файла: {e}")


def admin_broadcast(message):
    text = message.text

    admin_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    admin_btn1 = types.KeyboardButton('✅ Да, отправить')
    admin_btn2 = types.KeyboardButton('❌ Нет, отменить')
    admin_markup.row(admin_btn1, admin_btn2)

    bot.send_message(message.chat.id,
                     f"📢 <b>Превью рассылки:</b>\n\n{text}\n\n"
                     f"Отправить это сообщение всем пользователям?",
                     parse_mode='html', reply_markup=admin_markup)

    bot.register_next_step_handler(message, confirm_broadcast, text)


def confirm_broadcast(message, broadcast_text):
    if message.text == '✅ Да, отправить':
        sent_msg = bot.send_message(message.chat.id, "🔄 Начинаю рассылку...")
        success, failed = broadcast_message(broadcast_text)

        admin_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        admin_btn1 = types.KeyboardButton('📊 Загрузить таблицу')
        admin_btn2 = types.KeyboardButton('📢 Сделать рассылку')
        admin_btn3 = types.KeyboardButton('📨 Написать в чат')
        admin_btn4 = types.KeyboardButton('↩️ Назад в главное меню')
        admin_markup.row(admin_btn1, admin_btn2)
        admin_markup.row(admin_btn3, admin_btn4)

        bot.send_message(message.chat.id,
                         f"✅ <b>Рассылка завершена!</b>\n"
                         f"✅ Успешно отправлено: {success}\n"
                         f"❌ Не удалось отправить: {failed}",
                         parse_mode='html', reply_markup=admin_markup)
    else:
        admin_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        admin_btn1 = types.KeyboardButton('📊 Загрузить таблицу')
        admin_btn2 = types.KeyboardButton('📢 Сделать рассылку')
        admin_btn3 = types.KeyboardButton('📨 Написать в чат')
        admin_btn4 = types.KeyboardButton('↩️ Назад в главное меню')
        admin_markup.row(admin_btn1, admin_btn2)
        admin_markup.row(admin_btn3, admin_btn4)

        bot.send_message(message.chat.id, "❌ Рассылка отменена.", reply_markup=admin_markup)


def send_to_chat(message):
    try:
        parts = message.text.split(' ', 1)
        if len(parts) < 2:
            bot.send_message(message.chat.id, "❌ Неверный формат. Используйте: <chat_id> <сообщение>")
            return

        chat_id = parts[0]
        text = parts[1]

        if not chat_id.isdigit():
            bot.send_message(message.chat.id, "❌ Chat ID должен быть числом")
            return

        chat_id = int(chat_id)

        try:
            bot.send_message(chat_id, f"📨 Сообщение от администратора:\n\n{text}")
            bot.send_message(message.chat.id, f"✅ Сообщение успешно отправлено в чат {chat_id}")
        except Exception as e:
            bot.send_message(message.chat.id, f"❌ Не удалось отправить сообщение: {e}")
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Ошибка: {e}")

    admin_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    admin_btn1 = types.KeyboardButton('📊 Загрузить таблицу')
    admin_btn2 = types.KeyboardButton('📢 Сделать рассылку')
    admin_btn3 = types.KeyboardButton('📨 Написать в чат')
    admin_btn4 = types.KeyboardButton('↩️ Назад в главное меню')
    admin_markup.row(admin_btn1, admin_btn2)
    admin_markup.row(admin_btn3, admin_btn4)

    bot.send_message(message.chat.id, "Возвращаемся в админ панель...", reply_markup=admin_markup)


@auth_required
def find_row(message):
    save_user(message)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton('📋 Информация')
    btn2 = types.KeyboardButton('⚙️ Админ панель')
    btn3 = types.KeyboardButton('🔍 Поиск по номеру заказа')
    markup.row(btn1, btn2)
    markup.row(btn3)

    try:
        if not os.path.exists("Table.xlsx"):
            bot.send_message(message.chat.id, '❌ Таблица не загружена. Сначала загрузите таблицу.', reply_markup=markup)
            return

        table = load_workbook("Table.xlsx")
        lable = table.active
        row_p = True

        for row_num in range(1, lable.max_row + 1):
            cell_value = lable.cell(row=row_num, column=1).value
            if cell_value == message.text:
                row_p = False
                status = lable.cell(row=row_num, column=8).value if lable.cell(row=row_num, column=8).value else ""

                date_cell = lable.cell(row=row_num, column=9).value
                date_info = ""

                if status and "доставлен" in str(status).lower():
                    if date_cell:
                        try:
                            date_info = f"🟢 Был доставлен {date_cell.strftime('%d.%m.%Y')}"
                        except AttributeError:
                            date_info = f"🟢 Был доставлен {date_cell}"
                    else:
                        date_info = "🟢 Был доставлен (дата не указана)"
                elif status and "обработк" in str(status).lower():
                    date_info = "⏳ Дата поставки неизвестна"
                else:
                    if date_cell:
                        try:
                            date_info = f"📅 Планируемая дата поставки: {date_cell.strftime('%d.%m.%Y')}"
                        except AttributeError:
                            date_info = f"📅 Планируемая дата поставки: {date_cell}"
                    else:
                        date_info = "📅 Планируемая дата поставки не указана"

                bot.send_message(message.chat.id, f'''✅<b>Заказ №{message.text} найден в системе</b>:

📦<b>Наименование продукции</b>: {lable.cell(row=row_num, column=2).value}
🔢<b>Количество</b>: {lable.cell(row=row_num, column=4).value} {lable.cell(row=row_num, column=3).value}
💰<b>Цена за единицу</b>: {lable.cell(row=row_num, column=5).value} руб.
💵<b>Общая стоимость</b>: {lable.cell(row=row_num, column=6).value} руб.
🏢<b>Поставщик</b>: {lable.cell(row=row_num, column=7).value}
📊<b>Статус заказа</b>: {lable.cell(row=row_num, column=8).value}
{date_info}
                ''', parse_mode='html', reply_markup=markup)
                break

        if row_p:
            bot.send_message(message.chat.id, '❌ Заказ с указанным номером не найден в системе.', reply_markup=markup)
    except Exception as e:
        bot.send_message(message.chat.id, f'❌ Ошибка при поиске: {e}', reply_markup=markup)


@bot.message_handler(commands=['start'])
def start(message):
    init_users_db()
    save_user(message)

    user_id = message.from_user.id

    if user_id in authorized_users or check_user_authorization(user_id):
        if user_id not in authorized_users:
            authorized_users[user_id] = True

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton('📋 Информация')
        btn2 = types.KeyboardButton('⚙️ Админ панель')
        btn3 = types.KeyboardButton('🔍 Поиск по номеру заказа')
        markup.row(btn1, btn2)
        markup.row(btn3)

        bot.send_message(message.chat.id,
                         'Добро пожаловать в систему управления заказами Горно-Обогатительного Комбината. Выберите необходимую опцию:',
                         reply_markup=markup)
    else:
        request_password(message)


@bot.message_handler(func=lambda message: True)
@auth_required
def on_click(message):
    save_user(message)

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    btn1 = types.KeyboardButton('📋 Информация')
    btn2 = types.KeyboardButton('⚙️ Админ панель')
    btn3 = types.KeyboardButton('🔍 Поиск по номеру заказа')
    markup.row(btn1, btn2)
    markup.row(btn3)

    if message.text == "📋 Информация":
        bot.send_message(message.chat.id, '''📋 <b>СИСТЕМА УПРАВЛЕНИЯ ЗАКАЗАМИ ГГОК "ТРЕКСТЕР"</b>

Добро пожаловать в автоматизированную систему отслеживания заказов Горно-Обогатительного Комбината.

<b>Функциональные возможности системы:</b>

✅ <b>Отслеживание статуса заказа</b> — Получение актуальной информации о текущем этапе выполнения заказа
✅ <b>Поиск по номеру заказа</b> — Оперативное получение полной информации по присвоенному номеру ГОК
✅ <b>Архив заказов</b> — Доступ к историческим данным по предыдущим поставкам

Система обеспечивает получение информации в режиме реального времени и исключает необходимость обращения к менеджеру по стандартным запросам.

<b>Для работы с системой воспользуйтесь навигационным меню.</b>''', parse_mode='html', reply_markup=markup)

    elif message.text == '⚙️ Админ панель':
        password = bot.send_message(message.chat.id,
                                    '🔐 Для доступа к админ панели требуется авторизация. Введите пароль:')
        bot.register_next_step_handler(password, secure)

    elif message.text == '🔍 Поиск по номеру заказа':
        number = bot.send_message(message.chat.id,
                                  '🔍 Введите номер заказа для поиска\n(ГОК-{год_заказа}-{номер_заказа}):')
        bot.register_next_step_handler(number, find_row)

    elif message.text == '📊 Загрузить таблицу':
        msg = bot.send_message(message.chat.id, '📁 Прошу передать файл таблицы для загрузки.')
        bot.register_next_step_handler(msg, admin_table_add)

    elif message.text == '📢 Сделать рассылку':
        msg = bot.send_message(message.chat.id, '✍️ Введите текст для рассылки:')
        bot.register_next_step_handler(msg, admin_broadcast)

    elif message.text == '📨 Написать в чат':
        msg = bot.send_message(message.chat.id,
                               '💬 <b>Отправка сообщения в чат</b>\n\n'
                               'Введите данные в формате:\n'
                               '<code>&lt;chat_id&gt; &lt;текст сообщения&gt;</code>\n\n'
                               'Например:\n'
                               '<code>123456789 Привет! Это тестовое сообщение от администратора.</code>\n\n'
                               '💡 <b>Chat ID можно получить из файла users.xlsx</b>',
                               parse_mode='html')
        bot.register_next_step_handler(msg, send_to_chat)

    elif message.text == '↩️ Назад в главное меню':
        bot.send_message(message.chat.id, 'Возвращаемся в главное меню...', reply_markup=markup)


init_users_db()
bot.polling(none_stop=True)
